# -*- coding: utf-8 -*-
"""
Created on Fri Nov 29 13:07:12 2024

@author: willi
"""
import os
import glob
import numpy as np
import pandas as pd
import cv2
from matplotlib import pyplot as plt
#from sklearn.preprocessing import LabelEncoder
from shapely.geometry import LineString
import geopandas as gpd
#from matplotlib.patches import FancyBboxPatch
from matplotlib.lines import Line2D
from shapely.geometry import MultiPoint, Point

# find the intersection of cross section and coastal line

def shift_gdf(gdf, shift_x, shift_y):
    """
    Shift the geometry of a GeoDataFrame by a given distance in x and y directions.
    
    Parameters:
        gdf (GeoDataFrame): Input GeoDataFrame with CRS 3826 (TWD97).
        shift_x (float): Distance to shift along the x-axis (Easting).
        shift_y (float): Distance to shift along the y-axis (Northing).
    
    Returns:
        GeoDataFrame: GeoDataFrame with shifted geometries.
    """
    # Apply the shift to each geometry
    gdf['geometry'] = gdf['geometry'].apply(
        lambda geom: MultiPoint([Point(p.x + shift_x, shift_y - p.y) for p in geom.geoms])
        if geom.geom_type == 'MultiPoint' else geom
    )
    return gdf
def get_min_x_point(geometry):
    """
    Extracts the point with the minimum x-coordinate from a MultiPoint geometry.
    
    Parameters:
        geometry (shapely.geometry.MultiPoint or Point): Input geometry.
    
    Returns:
        Point or None: Point with the minimum x-coordinate, or None if geometry is invalid.
    """
    if geometry is None:  # 先檢查是否為 None
        return None
    if geometry.geom_type == 'MultiPoint':
        return min(geometry.geoms, key=lambda p: p.x)
    elif geometry.geom_type == 'Point':
        return geometry
    else:
        return None  # 處理其他無效的幾何類型



# read the cross section

CS = pd.read_excel('CrossSection_Reshaped_Extend_TWD97.xlsx')

y_starts = 30

SIZE_X = 896 
SIZE_Y = 512
n_classes=6 #Number of classes for segmentation


#Capture training image info as a list
Images = []
imfilenames = []
for directory_path in glob.glob("image/"): # change this folder name.
    for img_path in glob.glob(os.path.join(directory_path, "*.tif")):
        imfilenames.append(os.path.basename(img_path))
        img = cv2.imread(img_path) 
        # drop the last column and first row to make the num of column and row to 600 and 960
        cropped_img = img[1:, :-1]
        cropped_img = img[0:896, :512]
        Images.append(cropped_img)
#print(cropped_img.shape)       
#Convert list to array for machine learning processing        
Images = np.array(Images)
nI = Images.shape[0]
print(Images.shape)

    # upper left corner for TWD97.
# xul = 343627.902
# yul = 2770943.738

# NEW
xul = 343628.589
yul = 2770952.645
    
# 1. Load the pre-trained model
import keras  # Keras 3
def _dbg(x, *args, **kwargs):  # 舊模型裡 Lambda(_dbg) 的無害替身
    return x

model = keras.models.load_model(
    r'D:\OneDrive\桌面\GetCoastalLine.V2\UNET_20251011__253.h5',
    compile=False,
    custom_objects={'_dbg': _dbg},
    safe_mode=False  # 你信任自己的模型 → 關掉就能載
)

from openpyxl import load_workbook
output_dir = 'output_images/'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
excel_filename = 'output.xlsx'
# select the first image to do prediction
all_results = []
first_run = not os.path.exists(excel_filename)  # 判斷是否為第一次運行
for i in range(Images.shape[0]): # Change i to select which image to do the analysis

    input_data = Images[i,:,:,:]

    input_data = np.expand_dims(input_data, axis=0)

    prediction = model.predict(input_data)

    y_pred = np.argmax(prediction, axis=3)

    edges_pred = cv2.Canny((y_pred[0,:,:] == 1).astype(np.uint8), 0, 1)

    # 2. read the cross sections

    rgbImg = cv2.cvtColor(input_data[0,:,:,:], cv2.COLOR_BGR2RGB)

    ## Here is to get the intersection of cross sections and the predicted coastal line.

    indices = np.argwhere(edges_pred == 255)
    BdX = xul+indices[:,1]
    BdY = yul-indices[:,0]

    combined_array = np.vstack((BdX, BdY))
    # Create DataFrame from the combined array
    Bd = pd.DataFrame(combined_array).T
    Bd.columns = ['x', 'y']

    geometry = [Point(xy) for xy in zip(Bd['x'], Bd['y'])]

    # Convert to GeoDataFrame
    gdf = gpd.GeoDataFrame(Bd, geometry=geometry)

    Bd_line = LineString(gdf.geometry.tolist())
    # EPSG:3826 for TWD97
    gdf_line = gpd.GeoDataFrame(index=[0], geometry=[Bd_line],crs=3826)

    #plt.figure()
    #gdf_line.plot(color='red')
    #plt.show()
    Intersections = [None] * len(CS)  # 預設為 None
    IDs = [None] * len(CS)

    for j in range(len(CS)):
            cs = CS.iloc[j]
            line = LineString([(cs.X1, cs.Y1), (cs.X2, cs.Y2)])

            intersection = line.intersection(gdf_line.unary_union)

            # 保持與 CS 長度一致，交點為空時保持 None
            IDs[j] = cs.ID
            Intersections[j] = intersection if not intersection.is_empty else None

    output = gpd.GeoDataFrame({"ID": IDs, "geometry": Intersections}, crs="EPSG:3826")

        # 提取 `x` 和 `y` 坐標，對 `Point` 和 `MultiPoint` 進行處理
    output['geometry'] = output['geometry'].apply(get_min_x_point)
    output['x'] = output.geometry.x
    output['y'] = output.geometry.y
        # 檢查結果
    print(output[["ID", "geometry", "x", "y"]])

    # output = output.sort_values(by='x')
    output_df = pd.DataFrame([output['x'].values, output['y'].values], columns=IDs)
    
    
    # Here is to create a plot.
    # fig, ax = plt.subplots(figsize=(12, 8))
    # plt.imshow(rgbImg)

    # # 繪製橫斷線 (黑色虛線)
    # ax.plot([CS.X1 - xul, CS.X2 - xul], [yul - CS.Y1, yul - CS.Y2], 'k--')

    # # 繪製海岸線 (紅色)
    # ax.plot(output['x'] - xul, yul - output['y'], 'r-', label='Coastal Line')

    # # 繪製輪廓線 (藍色)
    # plt.contour(edges_pred, colors='blue', linewidths=0.5, linestyles='dotted')

    # # 繪製交點 (紅色點)
    # filtered_output = output.dropna(subset=['x', 'y'])
    # if not filtered_output.empty:
    #     ax.scatter(filtered_output['x'] - xul, yul - filtered_output['y'], 
    #             color='red', s=20, label='Intersections')
    # # 設定圖例
    # handles = [
    #     Line2D([0], [0], color='r', lw=2, label='Coastal Line'),  # 海岸線
    #     Line2D([0], [0], color='blue', linestyle='-', label='Boundary'),  # 邊界
    #     Line2D([0], [0], color='k', linestyle='--', lw=2, label='Cross section line'),  # 橫斷線 (黑色虛線)
    #     plt.scatter([], [], color='red', s=20, label='Intersections')  # 交點
    # ]
    # ax.legend(handles=handles, loc='upper right', fontsize=10, fancybox=True, framealpha=0.8)
    # # 繪圖設定
    # ax.set_aspect('equal', adjustable='box')
    # plt.xlim([0, 512])
    # plt.ylim([896, 0])
    # ax.set_title(f"{imfilenames[i]}")
    # # plt.show()
    # # 存檔 & 關閉
    # plt.savefig(os.path.join(output_dir, f"{imfilenames[i]}"))
    # plt.close()


    # if first_run and i == 0:
    #         output_df.insert(0, 'filename', imfilenames[i])
            
    #         # **第一次寫入時包含標題**
    #         output_df.to_excel(excel_filename, index=False, header=True)
    # else:
    #     output_df.insert(0, 'filename', imfilenames[i])  # 插入 filename 欄位
            
    #         # **檢查 Excel 是否已存在，追加寫入**
    #     with pd.ExcelWriter(excel_filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
    #         workbook = load_workbook(excel_filename)
    #         sheet = workbook.active
    #         startrow = sheet.max_row  # 確保新數據附加到舊數據後
    #         output_df.to_excel(writer, index=False, header=False, startrow=startrow)
