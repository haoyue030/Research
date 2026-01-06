# -*- coding: utf-8 -*-
"""
Created on Fri Nov 29 13:07:12 2024

@author: willi
"""
import os
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
import glob
import numpy as np
import pandas as pd
import cv2
from matplotlib import pyplot as plt
from matplotlib.colors import ListedColormap, BoundaryNorm
from matplotlib.cm import get_cmap
#from sklearn.preprocessing import LabelEncoder
from shapely.geometry import LineString
import geopandas as gpd
#from matplotlib.patches import FancyBboxPatch
from matplotlib.lines import Line2D
from shapely.geometry import MultiPoint, Point
from openpyxl import load_workbook

# 字型
plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei']
plt.rcParams['axes.unicode_minus'] = False

# === 分類圖色盤：與訓練顯示一致（viridis），離散化為 6 類 ===
NUM_CLASSES = 6
_base = get_cmap('viridis')
_colors = _base(np.linspace(0, 1, NUM_CLASSES))[:, :3]   # 取 RGB
DISCRETE_VIRIDIS = ListedColormap(_colors)
DISCRETE_NORM = BoundaryNorm(np.arange(-0.5, NUM_CLASSES + 0.5, 1), NUM_CLASSES)

# find the intersection of cross section and coastal line

def shift_gdf(gdf, shift_x, shift_y):
    """
    Shift the geometry of a GeoDataFrame by a given distance in x and y directions.
    CRS 3826 (TWD97).
    """
    gdf = gdf.copy()
    gdf['geometry'] = gdf['geometry'].apply(
        lambda geom: MultiPoint([Point(p.x + shift_x, shift_y - p.y) for p in geom.geoms])
        if geom.geom_type == 'MultiPoint' else geom
    )
    return gdf

def get_min_x_point(geometry):
    """Return Point with minimum x; None if invalid."""
    if geometry is None:
        return None
    if geometry.geom_type == 'MultiPoint':
        return min(geometry.geoms, key=lambda p: p.x)
    elif geometry.geom_type == 'Point':
        return geometry
    else:
        return None

# read the cross section
CS = pd.read_excel('CrossSection_Reshaped_Extend_TWD97.xlsx')

y_starts = 30
SIZE_X = 896
SIZE_Y = 512
n_classes = 6  # Number of classes for segmentation

# Capture training image info as a list
Images = []
imfilenames = []
for directory_path in glob.glob("errorimage/"):  # change this folder name if needed
    for img_path in glob.glob(os.path.join(directory_path, "*.tif")):
        imfilenames.append(os.path.basename(img_path))
        img = cv2.imread(img_path)
        # drop the last column and first row, then crop to 896x512
        cropped_img = img[1:, :-1]
        cropped_img = cropped_img[0:896, 0:512]  # 修正：以上一行結果為基準裁切
        Images.append(cropped_img)

Images = np.array(Images)
nI = Images.shape[0]
print(Images.shape)

# upper left corner for TWD97
xul = 343628.589
yul = 2770952.645

# 1. Load the pre-trained model
from keras.models import load_model  # type: ignore
model = load_model('UNET_20250409__253.h5')

# 輸出資料夾
overlay_dir = 'output_images'      # 疊圖（原邏輯）
map_dir     = 'output_unetmaps'    # 新增：兩聯圖（左原圖右分類）
os.makedirs(overlay_dir, exist_ok=True)
os.makedirs(map_dir, exist_ok=True)

excel_filename = 'output.xlsx'
first_run = not os.path.exists(excel_filename)

# select the first image to do prediction
all_results = []
for i in range(Images.shape[0]):  # Change i to select which image to do the analysis

    input_data = Images[i, :, :, :]
    input_data = np.expand_dims(input_data, axis=0)

    prediction = model.predict(input_data, verbose=0)
    y_pred = np.argmax(prediction, axis=3)            # (1, 896, 512)
    y_pred_lbl = y_pred[0, :, :]                      # (896, 512)

    edges_pred = cv2.Canny((y_pred_lbl == 1).astype(np.uint8), 0, 1)

    # 2. read the cross sections
    rgbImg = cv2.cvtColor(input_data[0, :, :, :], cv2.COLOR_BGR2RGB)

    # === 圖1：兩聯圖（左=原圖，右=UNET分類結果；viridis 離散 6 類） ===
    h, w = y_pred_lbl.shape
    panel_w = 3.4
    fig_w = 2 * panel_w
    fig_h = panel_w * (h / w)
    fig, axes = plt.subplots(1, 2, figsize=(fig_w, fig_h), dpi=200)

    # 左：原圖
    axes[0].imshow(rgbImg)
    axes[0].set_axis_off()
    axes[0].set_title('原始影像', fontsize=16)

    # 右：分類
    axes[1].imshow(y_pred_lbl, cmap=DISCRETE_VIRIDIS, norm=DISCRETE_NORM, interpolation='nearest')
    axes[1].set_axis_off()
    axes[1].set_title('UNET分類結果', fontsize=16)

    plt.tight_layout(pad=0.2)
    stem = os.path.splitext(imfilenames[i])[0]
    plt.savefig(os.path.join(map_dir, f'{stem}_orig_and_unet.png'),
                bbox_inches='tight', pad_inches=0.05)
    plt.close(fig)

    ## Here is to get the intersection of cross sections and the predicted coastal line.
    indices = np.argwhere(edges_pred == 255)
    BdX = xul + indices[:, 1]
    BdY = yul - indices[:, 0]

    combined_array = np.vstack((BdX, BdY))
    Bd = pd.DataFrame(combined_array).T
    Bd.columns = ['x', 'y']

    geometry = [Point(xy) for xy in zip(Bd['x'], Bd['y'])]

    # Convert to GeoDataFrame
    gdf = gpd.GeoDataFrame(Bd, geometry=geometry)

    Bd_line = LineString(gdf.geometry.tolist())
    # EPSG:3826 for TWD97
    gdf_line = gpd.GeoDataFrame(index=[0], geometry=[Bd_line], crs=3826)

    Intersections = [None] * len(CS)
    IDs = [None] * len(CS)

    for j in range(len(CS)):
        cs = CS.iloc[j]
        line = LineString([(cs.X1, cs.Y1), (cs.X2, cs.Y2)])
        intersection = line.intersection(gdf_line.unary_union)
        IDs[j] = cs.ID
        Intersections[j] = intersection if not intersection.is_empty else None

    output = gpd.GeoDataFrame({"ID": IDs, "geometry": Intersections}, crs="EPSG:3826")

    # 提取 x, y
    output['geometry'] = output['geometry'].apply(get_min_x_point)
    output['x'] = output.geometry.x
    output['y'] = output.geometry.y

    print(output[["ID", "geometry", "x", "y"]])

    # output = output.sort_values(by='x')
    output_df = pd.DataFrame([output['x'].values, output['y'].values], columns=IDs)

    # === 疊圖（原邏輯） ===
    fig, ax = plt.subplots(figsize=(12, 8))
    plt.imshow(rgbImg)

    # 橫斷線 (黑色虛線) —— 原邏輯一次性呼叫，若需逐條可改 for 迴圈
    ax.plot([CS.X1 - xul, CS.X2 - xul], [yul - CS.Y1, yul - CS.Y2], 'k--')

    # 海岸線 (紅色)
    ax.plot(output['x'] - xul, yul - output['y'], 'r-', label='Coastal Line')

    # 輪廓線 (藍色)
    plt.contour(edges_pred, colors='blue', linewidths=0.5, linestyles='dotted')

    # 交點 (紅色點)
    filtered_output = output.dropna(subset=['x', 'y'])
    if not filtered_output.empty:
        ax.scatter(filtered_output['x'] - xul, yul - filtered_output['y'],
                   color='red', s=20, label='Intersections')

    handles = [
        Line2D([0], [0], color='r', lw=2, label='Coastal Line'),
        Line2D([0], [0], color='blue', linestyle='-', label='Boundary'),
        Line2D([0], [0], color='k', linestyle='--', lw=2, label='Cross section line'),
        plt.scatter([], [], color='red', s=20, label='Intersections')
    ]
    ax.legend(handles=handles, loc='upper right', fontsize=10, fancybox=True, framealpha=0.8)

    ax.set_aspect('equal', adjustable='box')
    plt.xlim([0, 512])
    plt.ylim([896, 0])
    ax.set_title(f"{imfilenames[i]}")

    output_path = os.path.join(overlay_dir, f"{imfilenames[i]}")
    plt.savefig(output_path)
    plt.close()

    # === Excel 寫入 ===
    if first_run and i == 0:
        output_df.insert(0, 'filename', imfilenames[i])
        output_df.to_excel(excel_filename, index=False, header=True)
    else:
        output_df.insert(0, 'filename', imfilenames[i])
        with pd.ExcelWriter(excel_filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            workbook = load_workbook(excel_filename)
            sheet = workbook.active
            startrow = sheet.max_row
            output_df.to_excel(writer, index=False, header=False, startrow=startrow)
