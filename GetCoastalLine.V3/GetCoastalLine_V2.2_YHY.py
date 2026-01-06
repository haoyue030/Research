# -*- coding: utf-8 -*-
import os, glob
import numpy as np
import pandas as pd
import cv2
import geopandas as gpd
from matplotlib import pyplot as plt
from matplotlib.lines import Line2D
from shapely.geometry import LineString, Point, MultiPoint, MultiLineString
from shapely.ops import unary_union
from openpyxl import load_workbook
import keras  # Keras 3

# ====== 類別設定 ======
SAND_CLASS = 0          # 沙灘類別
ROCK_CLASS = 2          # 礁石類別（請依你的標註調整）

# ---------- 工具函式 ----------
def shift_gdf(gdf, shift_x, shift_y):
    """將 GeoDataFrame 幾何平移 (EPSG:3826)。"""
    def _shift_geom(geom):
        if geom.geom_type == "MultiPoint":
            return MultiPoint([Point(p.x + shift_x, p.y + shift_y) for p in geom.geoms])
        if geom.geom_type == "Point":
            return Point(geom.x + shift_x, geom.y + shift_y)
        return geom
    gdf = gdf.copy()
    gdf["geometry"] = gdf["geometry"].apply(_shift_geom)
    return gdf

def get_max_x_point(geometry):
    """回傳 x 最大的點；無點回傳 None。"""
    if geometry is None or geometry.is_empty:
        return None
    if geometry.geom_type == "Point":
        return geometry
    if geometry.geom_type == "MultiPoint":
        return max(geometry.geoms, key=lambda p: p.x)
    if geometry.geom_type == "GeometryCollection":
        pts = [g for g in geometry.geoms if g.geom_type == "Point"]
        return max(pts, key=lambda p: p.x) if pts else None
    try:
        coords = list(geometry.coords)
        return Point(max(coords, key=lambda xy: xy[0]))
    except Exception:
        return None

def clean_mask(binary_mask, open_ks=3, close_ks=5):
    """形態學開/閉運算，去小洞與毛邊。輸入/輸出皆為 0/1。"""
    m = binary_mask.astype(np.uint8)
    if open_ks and open_ks > 1:
        k = np.ones((open_ks, open_ks), np.uint8)
        m = cv2.morphologyEx(m, cv2.MORPH_OPEN, k, iterations=1)
    if close_ks and close_ks > 1:
        k = np.ones((close_ks, close_ks), np.uint8)
        m = cv2.morphologyEx(m, cv2.MORPH_CLOSE, k, iterations=1)
    return (m > 0).astype(np.uint8)

def mask_to_multilines(binary_mask, xul, yul, *, min_area=800, min_len=120, keep_largest=True):
    """
    將二值遮罩轉 MultiLineString，並濾掉太小的輪廓。
    min_area, min_len 單位為像素與像素長度；keep_largest=True 僅保留最大面積那條。
    """
    mask8 = (binary_mask.astype(np.uint8) * 255)
    res = cv2.findContours(mask8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
    cnts = res[0] if len(res) == 2 else res[1]

    lines, areas = [], []
    for c in cnts:
        if c is None or len(c) < 2:
            continue
        A = abs(cv2.contourArea(c))
        P = cv2.arcLength(c, True)
        if A < min_area or P < min_len:
            continue

        pts = c.squeeze()
        if pts.ndim != 2 or pts.shape[0] < 2:
            continue
        xs = xul + pts[:, 0]   # col -> Easting
        ys = yul - pts[:, 1]   # row -> Northing
        lines.append(LineString(np.c_[xs, ys]))
        areas.append(A)

    if not lines:
        return None
    if keep_largest:
        k = int(np.argmax(areas))
        return MultiLineString([lines[k]])
    return MultiLineString(lines)


# ---------- 讀資料 ----------
CS = pd.read_excel('CrossSection_Reshaped_Extend_TWD97.xlsx')  # 需含 ID, X1,Y1,X2,Y2

# 影像左上角 TWD97
xul = 343628.589
yul = 2770952.645

# 讀取影像
Images = []
imfilenames = []
for img_path in glob.glob(os.path.join("errorimage", "*.tif")):
    imfilenames.append(os.path.basename(img_path))
    img = cv2.imread(img_path)
    # 裁切：先去除首行與末列，再裁為 896x512
    crop = img[1:, :-1]
    crop = crop[0:896, 0:512]  # rows=896, cols=512
    Images.append(crop)
Images = np.array(Images)
print("Images shape:", Images.shape)

# 模型
def _dbg(x, *args, **kwargs):
    return x

model = keras.models.load_model(
    r'D:\OneDrive\桌面\GetCoastalLine.V2\UNET_20251014__253.h5',
    compile=False,
    custom_objects={'_dbg': _dbg},
    safe_mode=False
)

# 輸出
output_dir = 'output_images'
os.makedirs(output_dir, exist_ok=True)
excel_filename = 'output.xlsx'
first_run = not os.path.exists(excel_filename)

# ---------- 主流程 ----------
for i in range(Images.shape[0]):
    bgr = Images[i]
    rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)

    # 預測
    pred = model.predict(np.expand_dims(bgr, axis=0), verbose=0)
    y_pred = np.argmax(pred, axis=3)[0]  # (896,512)

    # 類別遮罩
    sand_mask = (y_pred == SAND_CLASS).astype(np.uint8)
    rock_mask = (y_pred == ROCK_CLASS).astype(np.uint8)

    # 先清理，再取輪廓 → 幾何
    sand_mask = clean_mask(sand_mask, open_ks=0, close_ks=3)
    rock_mask = clean_mask(rock_mask, open_ks=0, close_ks=3)

    sand_geom = mask_to_multilines(sand_mask, xul, yul, min_area=100, min_len=30, keep_largest=True)
    rock_geom = mask_to_multilines(rock_mask, xul, yul, min_area=100, min_len=30, keep_largest=True)

    # 提供繪圖所需的輪廓陣列（保持原本的 contour 用法）
    edges_pred = cv2.Canny(sand_mask.astype(np.uint8), 0, 1)

    # 遮罩→多折線，用於幾何交會
    sand_geom = mask_to_multilines(sand_mask, xul, yul)
    rock_geom = mask_to_multilines(rock_mask, xul, yul)

    # 與每條橫斷線求交；最後一條若無沙灘則改用礁石
    IDs, geoms = [], []
    last_idx = len(CS) - 1
    for idx, cs in CS.iterrows():
        line_cs = LineString([(cs.X1, cs.Y1), (cs.X2, cs.Y2)])
        inter = None if sand_geom is None else line_cs.intersection(sand_geom)

        # 備援：最後一個斷面且沙灘無交會 → 用礁石
        if (inter is None or inter.is_empty) and idx == last_idx and rock_geom is not None:
            inter = line_cs.intersection(rock_geom)

        IDs.append(cs.ID)
        geoms.append(inter if inter and not inter.is_empty else None)

    # 取 x 最大點
    out = gpd.GeoDataFrame({"ID": IDs, "geometry": geoms}, crs="EPSG:3826")
    out["geometry"] = out["geometry"].apply(get_max_x_point)
    out["x"] = out.geometry.x
    out["y"] = out.geometry.y

    # 與你既有繪圖/print 區塊相容的變數
    output = out.copy()
    rgbImg = rgb

    # ======= 下面這一段依你的要求完全不改 =======
    # 提取 `x` 和 `y` 坐標，對 `Point` 和 `MultiPoint` 進行處理
    output['geometry'] = output['geometry'].apply(get_max_x_point)
    output['x'] = output.geometry.x
    output['y'] = output.geometry.y
        # 檢查結果
    print(output[["ID", "geometry", "x", "y"]])

    # output = output.sort_values(by='x')
    output_df = pd.DataFrame([output['x'].values, output['y'].values], columns=IDs)
    
    # Here is to create a plot.
    fig, ax = plt.subplots(figsize=(12, 8))
    plt.imshow(rgbImg)

    # 繪製橫斷線 (黑色虛線)
    ax.plot([CS.X1 - xul, CS.X2 - xul], [yul - CS.Y1, yul - CS.Y2], 'k--')

    # 繪製海岸線 (紅色)
    ax.plot(output['x'] - xul, yul - output['y'], 'r-', label='Coastal Line')

    # 繪製輪廓線 (藍色)
    plt.contour(edges_pred, colors='blue', linewidths=0.5, linestyles='dotted')

    # 繪製交點 (紅色點)
    filtered_output = output.dropna(subset=['x', 'y'])
    if not filtered_output.empty:
        ax.scatter(filtered_output['x'] - xul, yul - filtered_output['y'], 
                color='red', s=20, label='Intersections')
    # 設定圖例
    handles = [
        Line2D([0], [0], color='r', lw=2, label='Coastal Line'),  # 海岸線
        Line2D([0], [0], color='blue', linestyle='-', label='Boundary'),  # 邊界
        Line2D([0], [0], color='k', linestyle='--', lw=2, label='Cross section line'),  # 橫斷線 (黑色虛線)
        plt.scatter([], [], color='red', s=20, label='Intersections')  # 交點
    ]
    ax.legend(handles=handles, loc='upper right', fontsize=10, fancybox=True, framealpha=0.8)
    # 繪圖設定
    ax.set_aspect('equal', adjustable='box')
    plt.xlim([0, 512])
    plt.ylim([896, 0])
    ax.set_title(f"{imfilenames[i]}")
    # plt.show()
    # 存檔 & 關閉
    plt.savefig(os.path.join(output_dir, f"{imfilenames[i]}"))
    plt.close()

    # ====== Excel 寫入（維持原本邏輯，移至迴圈內）======
    if first_run and i == 0:
        output_df.insert(0, 'filename', imfilenames[i])
        # **第一次寫入時包含標題**
        output_df.to_excel(excel_filename, index=False, header=True)
    else:
        output_df.insert(0, 'filename', imfilenames[i])  # 插入 filename 欄位
        # **檢查 Excel 是否已存在，追加寫入**
        with pd.ExcelWriter(excel_filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            workbook = load_workbook(excel_filename)
            sheet = workbook.active
            startrow = sheet.max_row  # 確保新數據附加到舊數據後
            output_df.to_excel(writer, index=False, header=False, startrow=startrow)
